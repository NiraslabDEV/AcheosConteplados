import pandas as pd
import os
from datetime import datetime
import glob
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

def encontrar_arquivos_recentes():
    """Encontra os arquivos mais recentes de cada site na pasta DADOS EXTRAIDOS."""
    pasta_dados = "DADOS EXTRAIDOS"
    if not os.path.exists(pasta_dados):
        print(f"Pasta {pasta_dados} não encontrada!")
        return {}
    
    # Padrões de arquivo para cada site
    padroes = {
        'venderseuconsorcio': 'extracao_venderseuconsorcio.xlsx',
        'cartascontempladas': 'extracao_cartascontempladas.xlsx',
        'consorciocontemplado': 'extracao_consorciocontemplado.xlsx',
        'contempladosp': 'extracao_SP_contemplados.xlsx',
        'consorciocontemplado_sp': 'consorcio_contemplado_base.xlsx'  # Adicionando o arquivo CS
    }
    
    arquivos_recentes = {}
    
    for site, arquivo in padroes.items():
        caminho_arquivo = os.path.join(pasta_dados, arquivo)
        if os.path.exists(caminho_arquivo):
            arquivos_recentes[site] = caminho_arquivo
            print(f"Arquivo encontrado para {site}: {arquivo}")
        else:
            print(f"Nenhum arquivo encontrado para {site}")
    
    return arquivos_recentes

def ler_e_padronizar_dados(arquivo, site):
    """Lê um arquivo Excel e padroniza suas colunas."""
    print(f"\nLendo dados de {site}...")
    try:
        df = pd.read_excel(arquivo)
    except Exception as e:
        print(f"Erro ao ler arquivo {arquivo}: {str(e)}")
        return None
    
    # Adiciona coluna de origem
    df['Origem'] = site
    
    # Garante que todas as colunas necessárias existem
    colunas_padrao = [
        'Código', 'Tipo', 'Valor da carta', 'Entrada', 'Total de Parcelas',
        'Consórcio', 'Status', 'Fluxo de Pagamento', 'Vencimento', 'Observações', 'Origem'
    ]
    
    for coluna in colunas_padrao:
        if coluna not in df.columns:
            df[coluna] = ''
    
    # Seleciona e ordena as colunas
    df = df[colunas_padrao]
    
    # Padroniza tipos de dados
    df['Total de Parcelas'] = pd.to_numeric(df['Total de Parcelas'], errors='coerce').fillna(0).astype(int)
    df['Valor da carta'] = pd.to_numeric(df['Valor da carta'].astype(str).str.replace('R$', '').str.replace('.', '').str.replace(',', '.').str.strip(), errors='coerce').fillna(0)
    df['Entrada'] = pd.to_numeric(df['Entrada'].astype(str).str.replace('R$', '').str.replace('.', '').str.replace(',', '.').str.strip(), errors='coerce').fillna(0)
    df['Fluxo de Pagamento'] = df['Fluxo de Pagamento'].fillna('')
    df['Observações'] = df['Observações'].fillna('')
    
    # Extrair informações sobre juros a partir do fluxo de pagamento
    df['Valor Parcela'] = df.apply(extrair_valor_parcela, axis=1)
    
    # Calcular juros implícitos (valor total a pagar - valor da carta)
    df['Valor Total a Pagar'] = df.apply(
        lambda row: row['Valor Parcela'] * row['Total de Parcelas'] if row['Valor Parcela'] > 0 and row['Total de Parcelas'] > 0 else 0,
        axis=1
    )
    
    df['Juros Implícitos'] = df.apply(
        lambda row: max(0, row['Valor Total a Pagar'] - row['Valor da carta']) if row['Valor Total a Pagar'] > 0 else 0,
        axis=1
    )
    
    # Calcular percentual de juros em relação ao valor da carta
    df['Percentual Juros'] = df.apply(
        lambda row: (row['Juros Implícitos'] / row['Valor da carta'] * 100) if row['Valor da carta'] > 0 else 0,
        axis=1
    )
    
    # Calcular razão de investimento aprimorada considerando juros
    df['Razão Investimento'] = df.apply(calcular_razao_investimento, axis=1)
    
    # Adiciona coluna de percentual de entrada
    df['Percentual Entrada'] = df.apply(lambda row:
        (row['Entrada'] / row['Valor da carta'] * 100) if row['Valor da carta'] > 0 else 0,
        axis=1
    )
    
    print(f"Total de registros de {site}: {len(df)}")
    return df

def extrair_valor_parcela(row):
    """Extrai o valor da parcela do fluxo de pagamento."""
    fluxo = row['Fluxo de Pagamento']
    if not fluxo or pd.isna(fluxo):
        return 0
    
    try:
        # Tentar extrair valores no formato "X x R$ Y,ZZ"
        import re
        matches = re.findall(r'(\d+)\s*x\s*R\$\s*([\d.,]+)', fluxo)
        
        if matches:
            # Pegar a primeira correspondência
            qtd_parcelas, valor_str = matches[0]
            # Limpar e converter o valor
            valor = float(valor_str.replace('.', '').replace(',', '.'))
            return valor
    except Exception as e:
        print(f"Erro ao extrair valor da parcela: {e}")
    
    return 0

def calcular_razao_investimento(row):
    """Calcula a razão de investimento considerando valor da carta, entrada e juros."""
    valor_carta = row['Valor da carta']
    entrada = row['Entrada']
    perc_juros = row['Percentual Juros']
    
    if entrada <= 0 or valor_carta <= 0:
        return 0
    
    # Base: razão entre valor da carta e entrada
    razao_base = valor_carta / entrada
    
    # Fator de ajuste para juros (quanto menor o percentual de juros, melhor)
    # Limitar juros a um valor máximo razoável para evitar distorções
    perc_juros_ajustado = min(perc_juros, 100.0)
    
    # O limite superior para juros é assumido como 40% (ajustável conforme necessário)
    MAX_JUROS = 40.0
    
    # Se juros for zero ou muito baixo, favorece ao máximo
    if perc_juros_ajustado <= 1.0:
        fator_juros = 1.5  # Bonificação para juros muito baixos
    else:
        # Quanto mais próximo de MAX_JUROS, mais penalizado
        # O fator varia de 1.5 (juros baixos) a 0.5 (juros altos)
        fator_juros = max(0.5, 1.5 - (perc_juros_ajustado / MAX_JUROS))
    
    # Razão final ponderada pelos juros
    razao_final = razao_base * fator_juros
    
    # Arredondar para duas casas decimais
    return round(razao_final, 2)

def formatar_excel(writer, df):
    """Aplica formatação ao arquivo Excel."""
    # Obter a planilha ativa
    workbook = writer.book
    worksheet = writer.sheets['Consolidado']
    
    # Configurar cabeçalho
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for idx, col in enumerate(df.columns):
        cell = worksheet.cell(row=1, column=idx+1)
        cell.fill = header_fill
        cell.font = header_font
        
        # Configurar largura e alinhamento para cada coluna
        if col == 'Fluxo de Pagamento':
            worksheet.column_dimensions[chr(65 + idx)].width = 50
            # Configurar quebra de linha para todas as células da coluna
            for cell in worksheet[chr(65 + idx)][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        elif col in ['Valor da carta', 'Entrada', 'Razão Investimento']:
            worksheet.column_dimensions[chr(65 + idx)].width = 20
            # Formatar células numéricas
            for cell in worksheet[chr(65 + idx)][1:]:
                cell.number_format = '#,##0.00'
        elif col == 'Percentual Entrada':
            worksheet.column_dimensions[chr(65 + idx)].width = 20
            # Formatar células de percentual
            for cell in worksheet[chr(65 + idx)][1:]:
                cell.number_format = '0.00%'
        else:
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(col)
            )
            worksheet.column_dimensions[chr(65 + idx)].width = max_length + 2

def consolidar_dados():
    """Função principal que consolida os dados de todos os extratores."""
    print("\n" + "="*50)
    print("Iniciando consolidação dos dados...")
    print("="*50)
    
    # Encontrar arquivos mais recentes
    arquivos = encontrar_arquivos_recentes()
    if not arquivos:
        print("\nNenhum arquivo encontrado para consolidar!")
        return
    
    # Lista para armazenar os DataFrames
    dfs = []
    
    # Ler e padronizar cada arquivo
    for site, arquivo in arquivos.items():
        try:
            df = ler_e_padronizar_dados(arquivo, site)
            if df is not None and not df.empty:
                dfs.append(df)
        except Exception as e:
            print(f"Erro ao processar arquivo de {site}: {str(e)}")
    
    if not dfs:
        print("\nNenhum dado foi carregado!")
        return
    
    # Concatenar todos os DataFrames
    print("\nConcatenando dados...")
    df_final = pd.concat(dfs, ignore_index=True)
    
    # Limitar valores extremos de juros para não distorcer a análise
    df_final['Percentual Juros'] = df_final['Percentual Juros'].clip(upper=100.0)
    
    # Ordenar por Razão Investimento (decrescente)
    df_final = df_final.sort_values('Razão Investimento', ascending=False)
    
    # Criar pasta de relatórios se não existir
    pasta_relatorios = "DADOS EXTRAIDOS"
    if not os.path.exists(pasta_relatorios):
        os.makedirs(pasta_relatorios)
    
    # Remover colunas indesejadas e colunas de cálculo intermediário antes de salvar
    colunas_internas = ['Vencimento', 'Observações', 'Origem', 'Valor Parcela', 'Valor Total a Pagar', 'Juros Implícitos', 'Percentual Juros']
    colunas_para_salvar = [col for col in df_final.columns if col not in colunas_internas]
    df_salvar = df_final[colunas_para_salvar]
    
    # Salvar arquivo consolidado
    data_hora = datetime.now().strftime("%Y%m%d_%H%M")
    arquivo_saida = os.path.join(pasta_relatorios, f'consolidado_{data_hora}.xlsx')
    
    print(f"\nSalvando arquivo consolidado: {arquivo_saida}")
    writer = pd.ExcelWriter(arquivo_saida, engine='openpyxl')
    df_salvar.to_excel(writer, index=False, sheet_name='Consolidado')
    
    # Aplicar formatação
    formatar_excel(writer, df_salvar)
    
    # Salvar e fechar
    writer.close()
    
    # Gerar estatísticas
    print("\n" + "="*50)
    print("ESTATÍSTICAS DO ARQUIVO CONSOLIDADO")
    print("="*50)
    print(f"\nTotal de registros: {len(df_final)}")
    
    print("\nRegistros por origem:")
    for origem in df_final['Origem'].unique():
        total = len(df_final[df_final['Origem'] == origem])
        print(f"- {origem}: {total} registros")
    
    print("\nRegistros por tipo:")
    for tipo in df_final['Tipo'].unique():
        total = len(df_final[df_final['Tipo'] == tipo])
        print(f"- {tipo}: {total} registros")
    
    print("\nRegistros por status:")
    for status in df_final['Status'].unique():
        total = len(df_final[df_final['Status'] == status])
        print(f"- {status}: {total} registros")
    
    print("\nEstatísticas de Juros:")
    # Considerar apenas registros com juros válidos
    df_juros_validos = df_final[df_final['Percentual Juros'] > 0]
    if len(df_juros_validos) > 0:
        juros_medios = df_juros_validos['Percentual Juros'].mean()
        juros_medianos = df_juros_validos['Percentual Juros'].median()
        juros_minimos = df_juros_validos['Percentual Juros'].min()
        juros_maximos = df_juros_validos['Percentual Juros'].max()
        print(f"- Juros Médios: {juros_medios:.2f}%")
        print(f"- Juros Medianos: {juros_medianos:.2f}%")
        print(f"- Juros Mínimos: {juros_minimos:.2f}%")
        print(f"- Juros Máximos: {juros_maximos:.2f}%")
        print(f"- Registros com juros calculados: {len(df_juros_validos)} de {len(df_final)}")
    else:
        print("- Nenhum registro com juros válidos encontrado")
    
    print("\nEstatísticas de Entrada:")
    entrada_media = df_final['Percentual Entrada'].mean()
    entrada_mediana = df_final['Percentual Entrada'].median()
    entrada_minima = df_final['Percentual Entrada'].min()
    entrada_maxima = df_final['Percentual Entrada'].max()
    print(f"- Percentual Médio de Entrada: {entrada_media:.2f}%")
    print(f"- Percentual Mediano de Entrada: {entrada_mediana:.2f}%")
    print(f"- Percentual Mínimo de Entrada: {entrada_minima:.2f}%")
    print(f"- Percentual Máximo de Entrada: {entrada_maxima:.2f}%")
    
    print("\nMelhores oportunidades (Top 5):")
    top_5 = df_final.head(5)
    for i, (_, row) in enumerate(top_5.iterrows(), 1):
        print(f"{i}. {row['Tipo']} - R$ {row['Valor da carta']:,.2f}")
        print(f"   Entrada: R$ {row['Entrada']:,.2f} ({row['Percentual Entrada']:.1f}%)")
        print(f"   Parcelas: {row['Total de Parcelas']} x R$ {row['Valor Parcela']:,.2f}")
        print(f"   Juros estimados: {row['Percentual Juros']:.2f}%")
        print(f"   Razão de investimento: {row['Razão Investimento']:.2f}")
        print()
    
    print(f"\nArquivo consolidado salvo com sucesso em: {arquivo_saida}")

if __name__ == "__main__":
    consolidar_dados() 