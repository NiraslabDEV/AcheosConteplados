from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import time
import re
import logging
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import platform
from webdriver_manager.chrome import ChromeDriverManager

# Configuração do sistema de logs
def configurar_logging():
    # Criar diretório de logs se não existir
    if not os.path.exists('logs'):
        os.makedirs('logs')
    
    # Nome do arquivo de log com timestamp
    log_filename = f'logs/extracao_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
    
    # Configurar o logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(log_filename),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

def configurar_driver():
    logger = logging.getLogger(__name__)
    logger.info("Iniciando configuração do driver Chrome")
    
    try:
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--window-size=1920,1080")
        
        # Configuração específica para Windows
        if platform.system() == 'Windows':
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
        else:
            driver = webdriver.Chrome(options=chrome_options)
            
        logger.info("Driver Chrome configurado com sucesso")
        return driver
    except Exception as e:
        logger.error(f"Erro ao configurar driver Chrome: {str(e)}")
        raise

def extrair_numero_parcelas(fluxo_pagamento):
    """Extrai o número de parcelas do fluxo de pagamento."""
    match = re.search(r'(\d+)\s*x', fluxo_pagamento)
    if match:
        return match.group(1)
    return ''

def padronizar_nome_consorcio(nome):
    """Padroniza o nome do consórcio."""
    nome = nome.strip()
    
    # Casos especiais
    if 'CAIXA' in nome.upper() or 'CNP' in nome.upper():
        return 'Caixa Econômica'
    if 'PORTO SEGURO' in nome.upper():
        return 'Porto Seguro'
    
    # Padronizar para primeira letra maiúscula
    palavras = nome.split()
    palavras_padronizadas = []
    
    for palavra in palavras:
        # Tratar palavras especiais
        if palavra.upper() in ['SA', 'S/A', 'LTDA', 'ME', 'EPP', 'EIRELI']:
            palavras_padronizadas.append(palavra.upper())
        else:
            palavras_padronizadas.append(palavra.capitalize())
    
    return ' '.join(palavras_padronizadas)

def extrair_dados_tabela(url):
    logger = logging.getLogger(__name__)
    logger.info(f"Iniciando extração de dados da URL: {url}")
    
    # Identificar o tipo baseado na URL
    tipo = "Imoveis" if "imoveis" in url.lower() else "Veiculos"
    logger.info(f"Tipo de consórcio identificado: {tipo}")
    
    driver = configurar_driver()
    dados = []
    
    try:
        logger.info("Acessando página...")
        driver.get(url)
        time.sleep(5)
        
        logger.info("Procurando tabela na página...")
        try:
            tabela = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table"))
            )
            logger.info("Tabela encontrada")
        except Exception as e:
            logger.error(f"Erro ao encontrar tabela: {str(e)}")
            logger.debug(f"HTML da página: {driver.page_source}")
            raise
        
        linhas = tabela.find_elements(By.TAG_NAME, "tr")
        logger.info(f"Encontradas {len(linhas)} linhas na tabela")
        
        # Validar estrutura da tabela
        if len(linhas) <= 1:
            logger.warning("Tabela vazia ou apenas com cabeçalho")
            return dados
        
        # Logging dos cabeçalhos para debug
        headers = [th.text for th in linhas[0].find_elements(By.TAG_NAME, "th")]
        logger.info(f"Cabeçalhos encontrados: {headers}")
        
        # Encontrar índice da coluna ID
        try:
            id_index = headers.index('ID')
            logger.info(f"Coluna ID encontrada no índice {id_index}")
        except ValueError:
            logger.error("Coluna ID não encontrada nos cabeçalhos")
            return dados
        
        for idx, linha in enumerate(linhas[1:], 1):
            logger.debug(f"Processando linha {idx}")
            try:
                # Tentar encontrar células tanto por td quanto por th
                colunas = linha.find_elements(By.TAG_NAME, "td")
                if not colunas:
                    colunas = linha.find_elements(By.TAG_NAME, "th")
                
                if len(colunas) >= 8:
                    # Log do conteúdo da linha para debug
                    conteudo_linha = [col.text.strip() for col in colunas]
                    logger.debug(f"Conteúdo da linha {idx}: {conteudo_linha}")
                    
                    # Tentar extrair ID de diferentes maneiras
                    id_cell = colunas[id_index]
                    id_texto = id_cell.text.strip()
                    
                    # Tentar encontrar o ID em diferentes formatos
                    id_patterns = [
                        r'CC(\d+)',  # Formato padrão CC seguido de números
                        r'(\d+)',    # Apenas números
                        r'[Cc][Cc]\s*(\d+)',  # CC ou cc com possíveis espaços
                        r'[Cc][Cc][-_]?(\d+)'  # CC ou cc com possível hífen ou underscore
                    ]
                    
                    id_consorcio = None
                    for pattern in id_patterns:
                        match = re.search(pattern, id_texto)
                        if match:
                            id_consorcio = f"CT{match.group(1)}"
                            break
                    
                    if not id_consorcio:
                        logger.warning(f"ID não encontrado na linha {idx}. Conteúdo da célula: '{id_texto}'")
                        # Tentar buscar em elementos filhos
                        for elem in id_cell.find_elements(By.XPATH, ".//*"):
                            elem_text = elem.text.strip()
                            for pattern in id_patterns:
                                match = re.search(pattern, elem_text)
                                if match:
                                    id_consorcio = f"CT{match.group(1)}"
                                    break
                            if id_consorcio:
                                break
                    
                    if not id_consorcio:
                        logger.warning(f"Não foi possível extrair ID da linha {idx}")
                        continue
                    
                    # Extrair demais campos
                    try:
                        credito = colunas[headers.index('Crédito')].text.strip()
                        entrada = colunas[headers.index('Entrada')].text.strip()
                        divida = colunas[headers.index('Dívida')].text.strip()
                        consorcio = colunas[headers.index('Consórcio')].text.strip()
                        consorcio = padronizar_nome_consorcio(consorcio)
                        status = colunas[headers.index('Status')].text.strip()
                        vencimento = colunas[headers.index('Vencimento')].text.strip()
                        observacoes = colunas[headers.index('Observações')].text.strip().rstrip('.')
                        
                        # Filtrar registros da Caixa Econômica
                        if 'CAIXA' in consorcio.upper() or 'CNP' in consorcio.upper():
                            logger.info(f"Registro {id_consorcio} da Caixa Econômica será ignorado")
                            continue
                        
                        # Calcular entrada total (entrada + 5%)
                        entrada = calcular_entrada(credito, entrada)
                        
                        # Extrair número de parcelas do fluxo de pagamento
                        total_parcelas = extrair_numero_parcelas(divida)
                        
                        # Processar fluxo de pagamento com base nas observações
                        fluxo_pagamento = divida.rstrip('.')
                        if observacoes:
                            # Padrões para diferentes formatos de observações
                            padroes = [
                                r"As (\d+) primeiras parcelas serão de R\$ ([0-9,.]+)",
                                r"Sendo as (\d+) primeiras parcelas de R\$ ([0-9,.]+)",
                                r"As (\d+) primeiras parcelas de R\$ ([0-9,.]+)"
                            ]
                            
                            for padrao in padroes:
                                match = re.search(padrao, observacoes)
                                if match:
                                    num_parcelas_diferentes = int(match.group(1))
                                    valor_parcelas_diferentes = match.group(2)
                                    
                                    # Extrair valor padrão das parcelas
                                    padrao_valor_padrao = r"(\d+)\s*x\s*R\$\s*([0-9,.]+)"
                                    match_valor = re.search(padrao_valor_padrao, divida)
                                    if match_valor:
                                        total_parcelas = int(match_valor.group(1))
                                        valor_padrao = match_valor.group(2)
                                        
                                        # Calcular número de parcelas padrão
                                        parcelas_padrao = total_parcelas - num_parcelas_diferentes
                                        
                                        # Montar fluxo de pagamento com quebra de linha
                                        fluxo_pagamento = f"{num_parcelas_diferentes} x R$ {valor_parcelas_diferentes}\n{parcelas_padrao} x R$ {valor_padrao}"
                                        break
                        
                        # Validar campos obrigatórios
                        if not all([credito, entrada, divida, consorcio, status]):
                            logger.warning(f"Campos obrigatórios ausentes na linha {idx}")
                            continue
                        
                        dados.append({
                            'Código': id_consorcio,
                            'Tipo': tipo,
                            'Valor da carta': credito,
                            'Entrada': entrada,
                            'Total de Parcelas': total_parcelas,
                            'Consórcio': consorcio,
                            'Status': status,
                            'Fluxo de Pagamento': fluxo_pagamento,
                            'Vencimento': vencimento,
                            'Observações': observacoes
                        })
                        logger.info(f"Linha {idx} processada com sucesso: ID {id_consorcio}")
                    except ValueError as e:
                        logger.error(f"Erro ao encontrar coluna na linha {idx}: {str(e)}")
                        continue
                    
            except Exception as e:
                logger.error(f"Erro ao processar linha {idx}: {str(e)}")
                continue
    
    except Exception as e:
        logger.error(f"Erro ao processar página: {str(e)}")
    
    finally:
        logger.info(f"Finalizando extração. Total de registros extraídos: {len(dados)}")
        driver.quit()
    
    return dados

def formatar_excel(arquivo):
    """Formata o arquivo Excel com colunas mais largas e conteúdo centralizado."""
    logger = logging.getLogger(__name__)
    logger.info("Iniciando formatação do arquivo Excel")
    
    try:
        # Carregar o workbook
        wb = load_workbook(arquivo)
        ws = wb.active
        
        # Definir largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            # Encontrar a largura máxima na coluna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Ajustar largura da coluna (com um mínimo de 15)
            adjusted_width = max(max_length + 2, 15)
            ws.column_dimensions[column].width = adjusted_width
        
        # Formatar valores monetários e centralizar conteúdo
        colunas_monetarias = ['Valor da carta', 'Entrada']
        for row in ws.rows:
            for cell in row:
                # Centralizar todo o conteúdo
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Formatar valores monetários
                if ws.cell(row=1, column=cell.column).value in colunas_monetarias:
                    if cell.row > 1 and cell.value:  # Pular cabeçalho e células vazias
                        valor = str(cell.value)
                        # Remover R$ e pontos, manter vírgula
                        valor = valor.replace('R$', '').replace('.', '').strip()
                        cell.value = valor
                
                # Configurar quebra de linha para a coluna Fluxo de Pagamento
                if ws.cell(row=1, column=cell.column).value == 'Fluxo de Pagamento':
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    # Aumentar a altura da linha para acomodar o conteúdo
                    ws.row_dimensions[cell.row].height = 30
        
        # Salvar as alterações
        wb.save(arquivo)
        logger.info("Formatação do arquivo Excel concluída com sucesso")
        
    except Exception as e:
        logger.error(f"Erro ao formatar arquivo Excel: {str(e)}")

def carregar_dados_existentes(arquivo):
    """Carrega os dados do arquivo Excel existente."""
    logger = logging.getLogger(__name__)
    try:
        if os.path.exists(arquivo):
            logger.info(f"Carregando dados existentes do arquivo {arquivo}")
            df_existente = pd.read_excel(arquivo)
            return df_existente
        else:
            logger.info("Arquivo não encontrado. Será criado um novo arquivo.")
            return pd.DataFrame()
    except Exception as e:
        logger.error(f"Erro ao carregar arquivo existente: {str(e)}")
        return pd.DataFrame()

def atualizar_dados(dados_novos, dados_existentes):
    """
    Atualiza os dados existentes com os novos dados.
    Remove registros vendidos e atualiza status dos reservados.
    """
    logger = logging.getLogger(__name__)
    
    if dados_existentes.empty:
        logger.info("Não há dados existentes. Retornando dados novos.")
        return pd.DataFrame(dados_novos)
    
    # Converter listas para DataFrames
    df_novos = pd.DataFrame(dados_novos)
    
    # Criar conjunto de IDs existentes para busca rápida
    ids_existentes = set(dados_existentes['Código'].values)
    
    # Lista para armazenar registros atualizados
    registros_atualizados = []
    
    # Processar novos dados
    for _, row_novo in df_novos.iterrows():
        codigo = row_novo['Código']
        status = row_novo['Status']
        
        # Se o registro existe, verificar mudanças
        if codigo in ids_existentes:
            registro_existente = dados_existentes[dados_existentes['Código'] == codigo].iloc[0]
            
            # Se o status mudou para "Vendido", não incluir no novo DataFrame
            if status.upper() == 'VENDIDO':
                logger.info(f"Registro {codigo} foi vendido e será removido")
                continue
            
            # Comparar valores dos campos relevantes
            campos_comparacao = ['Valor da carta', 'Entrada', 'Total de Parcelas', 'Status', 'Fluxo de Pagamento', 'Vencimento']
            houve_alteracao = False
            
            for campo in campos_comparacao:
                if str(registro_existente[campo]) != str(row_novo[campo]):
                    logger.info(f"Campo {campo} do registro {codigo} foi alterado de '{registro_existente[campo]}' para '{row_novo[campo]}'")
                    houve_alteracao = True
            
            if houve_alteracao:
                logger.info(f"Registro {codigo} foi atualizado")
                registros_atualizados.append(row_novo.to_dict())
            else:
                # Manter registro existente se não houve alterações
                registros_atualizados.append(registro_existente.to_dict())
        else:
            # Novo registro, adicionar apenas se não estiver vendido
            if status.upper() != 'VENDIDO':
                logger.info(f"Novo registro adicionado: {codigo}")
                registros_atualizados.append(row_novo.to_dict())
    
    # Criar novo DataFrame com os registros atualizados
    df_atualizado = pd.DataFrame(registros_atualizados)
    
    # Ordenar por código para manter consistência
    if not df_atualizado.empty:
        df_atualizado = df_atualizado.sort_values('Código')
    
    return df_atualizado

def calcular_entrada(valor_carta, entrada_atual):
    """
    Calcula a entrada total (entrada + 5% do valor da carta)
    """
    try:
        # Remover R$ e converter para float
        valor_carta = float(valor_carta.replace('R$', '').replace('.', '').replace(',', '.').strip())
        entrada_atual = float(entrada_atual.replace('R$', '').replace('.', '').replace(',', '.').strip())
        
        # Calcular 5% do valor da carta
        cinco_porcento = valor_carta * 0.05
        
        # Calcular entrada total
        entrada_total = entrada_atual + cinco_porcento
        
        # Formatar o valor para exibição
        return f"R$ {entrada_total:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
    except Exception as e:
        logging.error(f"Erro ao calcular entrada: {str(e)}")
        return entrada_atual

def processar_linha(linha, tipo, codigos_existentes):
    try:
        colunas = linha.find_elements(By.TAG_NAME, "td")
        if len(colunas) < 9:
            return None

        id_consorcio = colunas[1].text.strip()
        valor_carta = colunas[2].text.strip()
        entrada = colunas[3].text.strip()
        divida = colunas[4].text.strip()
        consorcio = colunas[5].text.strip()
        status = colunas[6].text.strip()
        vencimento = colunas[7].text.strip()
        observacoes = colunas[8].text.strip().rstrip('.')

        codigo = gerar_codigo_unico(codigos_existentes)
        valor_carta = formatar_valor_monetario(valor_carta)
        
        # Calcular entrada total (entrada + 5%)
        entrada = calcular_entrada(valor_carta, entrada)

        # Processa o fluxo de pagamento
        fluxo_pagamento = divida.rstrip('.')
        
        # Padrões para diferentes formatos de observações sobre parcelas iniciais
        padroes = [
            r"As (\d+) primeiras parcelas serão de R\$ ([0-9,.]+)",
            r"Sendo as (\d+) primeiras parcelas de R\$ ([0-9,.]+)",
            r"As (\d+) primeiras parcelas de R\$ ([0-9,.]+)"
        ]
        
        for padrao in padroes:
            match = re.search(padrao, observacoes)
            if match:
                num_parcelas_iniciais = int(match.group(1))
                valor_parcela_inicial = float(match.group(2).replace('.', '').replace(',', '.'))
                
                # Extrai o número total de parcelas
                total_parcelas = extrair_numero_parcelas(divida)
                if total_parcelas:
                    # Calcula o valor das parcelas restantes
                    valor_total = float(valor_carta.replace('R$ ', '').replace('.', '').replace(',', '.'))
                    entrada_valor = float(entrada.replace('R$ ', '').replace('.', '').replace(',', '.'))
                    valor_restante = valor_total - entrada_valor - (num_parcelas_iniciais * valor_parcela_inicial)
                    parcelas_restantes = total_parcelas - num_parcelas_iniciais
                    if parcelas_restantes > 0:
                        valor_parcela_restante = valor_restante / parcelas_restantes
                        # Construir o fluxo de pagamento com quebra de linha
                        fluxo_pagamento = f"{num_parcelas_iniciais} x R$ {valor_parcela_inicial:.2f}\n{parcelas_restantes} x R$ {valor_parcela_restante:.2f}"
                    else:
                        fluxo_pagamento = f"{num_parcelas_iniciais} x R$ {valor_parcela_inicial:.2f}"
                break

        return {
            'Código': codigo,
            'Tipo': tipo,
            'Valor da carta': valor_carta,
            'Entrada': entrada,
            'Total de Parcelas': extrair_numero_parcelas(divida),
            'Consórcio': consorcio,
            'Status': status,
            'Fluxo de Pagamento': fluxo_pagamento,
            'Vencimento': vencimento,
            'Observações': observacoes
        }
    except Exception as e:
        logging.error(f"Erro ao processar linha: {str(e)}")
        return None

def main():
    logger = configurar_logging()
    logger.info("Iniciando processo de extração de dados")
    
    # Cria pasta para os arquivos se não existir
    pasta_dados = "DADOS EXTRAIDOS"
    if not os.path.exists(pasta_dados):
        os.makedirs(pasta_dados)
    
    # Nome do arquivo base
    site_nome = "consorciocontemplado"
    arquivo_base = os.path.join(pasta_dados, f'extracao_{site_nome}.xlsx')
    
    # Carregar dados existentes
    dados_existentes = carregar_dados_existentes(arquivo_base)
    
    urls = [
        "https://www.consorciocontemplado.com.br/carta-de-automoveis-contempladas-para-venda.php",
        "https://www.consorciocontemplado.com.br/carta-de-imoveis-contempladas-para-venda.php"
    ]
    
    todos_dados = []
    
    for url in urls:
        logger.info(f"Processando URL: {url}")
        dados = extrair_dados_tabela(url)
        todos_dados.extend(dados)
        logger.info(f"Dados extraídos da URL: {len(dados)} registros")
        time.sleep(2)
    
    if todos_dados:
        logger.info(f"Total de registros coletados: {len(todos_dados)}")
        
        # Atualizar dados existentes com os novos dados
        df_atualizado = atualizar_dados(todos_dados, dados_existentes)
        
        # Log da estrutura do DataFrame
        logger.info(f"Estrutura do DataFrame atualizado:\n{df_atualizado.info()}")
        logger.info(f"Primeiras linhas do DataFrame atualizado:\n{df_atualizado.head()}")
        
        # Salvar arquivo
        df_atualizado.to_excel(arquivo_base, index=False)
        logger.info(f"Dados atualizados exportados com sucesso para {arquivo_base}")
        
        # Formatar o arquivo Excel
        formatar_excel(arquivo_base)
    else:
        logger.error("Nenhum dado foi extraído")

if __name__ == "__main__":
    main() 