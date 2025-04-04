#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import pandas as pd
import re
import time
import requests
from urllib3.exceptions import InsecureRequestWarning
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import glob
import platform
from webdriver_manager.chrome import ChromeDriverManager

# Desabilitar avisos de SSL
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

def formatar_moeda(valor):
    """Formata um valor numérico para o formato R$ X.XXX,XX"""
    return f'R$ {valor:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')

def formatar_numero(valor):
    """Formata um valor numérico apenas com vírgula (sem R$ e pontos)"""
    return f'{valor:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '')

def verificar_site_acessivel():
    """Verifica se o site está acessível antes de tentar o scraping."""
    try:
        response = requests.get('https://www.contempladosp.com.br', verify=False, timeout=10)
        return response.status_code == 200
    except Exception as e:
        print(f"Erro ao verificar acessibilidade do site: {str(e)}")
        return False

def configurar_driver():
    """Configura e retorna o driver do Selenium em modo headless."""
    print("Configurando o driver do Chrome...")
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--ignore-ssl-errors")
    chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    try:
        # Configuração específica para Windows
        if platform.system() == 'Windows':
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
        else:
            driver = webdriver.Chrome(options=chrome_options)
            
        driver.set_page_load_timeout(120)  # Aumentado para 120 segundos
        driver.implicitly_wait(30)  # Espera implícita de 30 segundos
        return driver
    except Exception as e:
        print(f"Erro ao configurar o driver: {str(e)}")
        raise

def normalizar_valor(valor_str):
    """Remove R$, pontos e vírgulas, convertendo para float."""
    if not valor_str:
        return 0.0
    try:
        return float(re.sub(r'[^\d,]', '', valor_str).replace(',', '.'))
    except ValueError:
        print(f"Erro ao normalizar valor: {valor_str}")
        return 0.0

def gerar_codigo(contador):
    """Gera um código único no formato SPXXXX, onde XXXX é um número sequencial"""
    return f"SP{contador:04d}"  # Formato SP0001, SP0002, etc.

def processar_parcelas(parcelas_str):
    """Extrai quantidade de parcelas e valor da parcela."""
    if not parcelas_str:
        return 0, 0.0, ""
    
    try:
        # Normaliza o X para minúsculo e adiciona R$ aos valores que não têm
        parcelas_str = parcelas_str.replace(" X ", " x ")
        parcelas_str = re.sub(r'(\d+)\s*x\s*(?!R\$)(\d)', r'\1 x R$ \2', parcelas_str)
        
        # Procura por padrões de múltiplas parcelas (ex: "sendo 4 x R$ 2.491,00 + 58 x R$ 1.934,00")
        if "sendo" in parcelas_str.lower():
            padrao = r'(\d+)\s*x\s*R\$\s*([\d.,]+)'
            matches = list(re.finditer(padrao, parcelas_str))
            
            total_parcelas = 0
            valor_ultima_parcela = 0.0
            detalhes_parcelas = []
            
            for match in matches:
                num = int(match.group(1))
                valor = normalizar_valor(match.group(2))
                total_parcelas += num
                valor_ultima_parcela = valor
                detalhes_parcelas.append(f"{num} x R$ {formatar_numero(valor)}")
            
            if total_parcelas > 0:
                observacao = "\n".join(detalhes_parcelas)
                return total_parcelas, valor_ultima_parcela, observacao
        
        # Procura por padrão simples com ou sem R$ (ex: "171 x 998,00" ou "171 x R$ 998,00")
        padrao_simples = r'(\d+)\s*x\s*(R\$\s*)?([\d.,]+)'
        match = re.search(padrao_simples, parcelas_str)
        if match:
            num_parcelas = int(match.group(1))
            valor_parcela = normalizar_valor(match.group(3))
            observacao = f"{num_parcelas} x R$ {formatar_numero(valor_parcela)}"
            return num_parcelas, valor_parcela, observacao
            
        # Se não encontrou nenhum padrão mas tem números e 'x', tenta extrair o total
        if any(c.isdigit() for c in parcelas_str):
            padrao_numeros = r'(\d+)\s*[xX]'
            matches = list(re.finditer(padrao_numeros, parcelas_str))
            if matches:
                total_parcelas = sum(int(match.group(1)) for match in matches)
                # Tenta extrair o valor
                valor_match = re.search(r'[xX]\s*(R\$\s*)?([\d.,]+)', parcelas_str)
                if valor_match:
                    valor = normalizar_valor(valor_match.group(2))
                    observacao = f"{total_parcelas} x R$ {formatar_numero(valor)}"
                    return total_parcelas, valor, observacao
                return total_parcelas, 0.0, parcelas_str
            
    except Exception as e:
        print(f"Erro ao processar parcelas: {parcelas_str}")
        print(f"Erro: {str(e)}")
    
    # Se não conseguiu extrair, retorna os valores padrão
    return 0, 0.0, parcelas_str

def normalizar_administradora(nome):
    """Normaliza o nome da administradora."""
    nome_upper = nome.upper()
    if "CAIXA XS5" in nome_upper or "CAIXA CNP" in nome_upper:
        return "Caixa Econômica"
    return nome

def extrair_dados_tabela(soup, tipo, contador_inicial=1):
    """Extrai dados de uma tabela específica."""
    print(f"\nExtraindo dados da tabela de {tipo}...")
    dados = []
    contador = contador_inicial
    
    # Procura pela tabela principal que contém os dados
    tabela = soup.find('table', {'class': 'table'})
    if not tabela:
        print(f"Não foi possível encontrar a tabela de {tipo}")
        return dados, contador
    
    linhas = tabela.find_all('tr')[1:]  # Pula o cabeçalho
    print(f"Encontradas {len(linhas)} linhas na tabela de {tipo}")
    
    for linha in linhas:
        colunas = linha.find_all('td')
        if len(colunas) >= 4:
            try:
                # Verifica o status (se disponível)
                status = "Disponível"
                if len(colunas) > 4:  # Se tiver coluna de status
                    status_original = colunas[4].text.strip().lower()
                    if "vendida" in status_original:
                        continue  # Ignora cotas vendidas
                    elif "reservada" in status_original:
                        status = "Indisponível"
                
                # Extrai os dados básicos
                credito = colunas[0].text.strip()
                entrada = colunas[1].text.strip()
                parcelas = colunas[2].text.strip()
                administradora = normalizar_administradora(colunas[3].text.strip())
                
                # Ignora cotas das administradoras especificadas
                if ("YAMAHA" in administradora.upper() or 
                    "RODOBENS" in administradora.upper() or 
                    "SICREDI" in administradora.upper() or
                    "ANCORA" in administradora.upper() or
                    "CAIXA" in administradora.upper() or
                    "HS" in administradora.upper() or
                    "UNICOOB" in administradora.upper() or
                    "EMBRACON" in administradora.upper() or
                    "VOLKSWAGEN" in administradora.upper()):
                    continue
                
                # Verifica se há observações específicas na linha
                obs_original = ""
                if len(colunas) > 5:
                    obs_original = colunas[5].text.strip()
                
                # Processa os valores
                credito_valor = normalizar_valor(credito)
                entrada_valor = normalizar_valor(entrada)
                
                # Calcula 5% do valor total do crédito e soma à entrada
                cinco_por_cento_credito = credito_valor * 0.05
                entrada_valor = entrada_valor + cinco_por_cento_credito
                
                # Se houver observação com detalhamento de parcelas, usa ela
                if "sendo" in obs_original.lower():
                    num_parcelas, valor_parcela, obs_parcelas = processar_parcelas(obs_original)
                else:
                    num_parcelas, valor_parcela, obs_parcelas = processar_parcelas(parcelas)
                
                # Se não conseguiu extrair observação das parcelas, usa o texto original
                if not obs_parcelas and obs_original:
                    obs_parcelas = obs_original
                
                # Se o número de parcelas for zero, tenta calcular a partir do fluxo de pagamento
                if num_parcelas == 0 and obs_parcelas:
                    padrao = r'(\d+)\s*x'
                    matches = re.finditer(padrao, obs_parcelas)
                    total_parcelas = sum(int(match.group(1)) for match in matches)
                    if total_parcelas > 0:
                        num_parcelas = total_parcelas
                
                # Formata o fluxo de pagamento substituindo + por quebra de linha
                if obs_parcelas:
                    obs_parcelas = obs_parcelas.replace(" + ", "\n").rstrip('.')
                
                # Extrai observações da cota (coluna 5)
                observacoes = obs_original if obs_original else ""
                
                dados.append({
                    'Código': gerar_codigo(contador),
                    'Tipo': tipo,
                    'Valor da carta': formatar_numero(credito_valor),
                    'Entrada': formatar_numero(entrada_valor),
                    'Total de Parcelas': num_parcelas,
                    'Consórcio': administradora,
                    'Status': status,
                    'Fluxo de Pagamento': obs_parcelas,
                    'Vencimento': (datetime.now() + pd.Timedelta(days=10)).strftime('%d/%m/%Y'),
                    'Observações': observacoes
                })
                contador += 1
            except Exception as e:
                print(f"Erro ao processar linha: {str(e)}")
                continue
    
    print(f"Extraídos {len(dados)} registros da tabela de {tipo}")
    return dados, contador

def criar_dataframe(dados_imoveis, dados_veiculos):
    """Cria um DataFrame combinando dados de imóveis e veículos."""
    print("\nCriando DataFrame com os dados extraídos...")
    todos_dados = dados_imoveis + dados_veiculos
    df = pd.DataFrame(todos_dados)
    
    # Definir a ordem das colunas
    colunas = ['Código', 'Tipo', 'Valor da carta', 'Entrada', 'Total de Parcelas', 
               'Consórcio', 'Status', 'Fluxo de Pagamento', 'Vencimento', 'Observações']
    df = df[colunas]
    
    # Garantir tipos corretos
    df['Total de Parcelas'] = pd.to_numeric(df['Total de Parcelas'], errors='coerce').fillna(0).astype(int)
    df['Valor da carta'] = df['Valor da carta'].fillna('0')
    df['Entrada'] = df['Entrada'].fillna('0')
    df['Fluxo de Pagamento'] = df['Fluxo de Pagamento'].fillna('')
    df['Observações'] = df['Observações'].fillna('')
    
    print(f"DataFrame criado com {len(df)} registros")
    return df

def salvar_excel(df, data):
    """Salva o DataFrame em um arquivo Excel."""
    print("\nSalvando arquivo Excel...")
    pasta_dados = 'DADOS EXTRAIDOS'
    if not os.path.exists(pasta_dados):
        os.makedirs(pasta_dados)
    
    # Nome do arquivo
    nome_arquivo = 'extracao_SP_contemplados.xlsx'
    caminho_completo = os.path.join(pasta_dados, nome_arquivo)
    
    # Remove o arquivo antigo se existir
    if os.path.exists(caminho_completo):
        os.remove(caminho_completo)
    
    # Salvar primeiro como Excel
    df.to_excel(caminho_completo, index=False)
    
    # Abrir o arquivo com openpyxl para formatação
    from openpyxl import load_workbook
    workbook = load_workbook(caminho_completo)
    worksheet = workbook.active
    
    # Definir larguras específicas para cada coluna (valores mais conservadores)
    larguras_colunas = {
        'A': 12,  # Código
        'B': 15,  # Tipo
        'C': 20,  # Valor da carta
        'D': 20,  # Entrada
        'E': 15,  # Total de Parcelas
        'F': 25,  # Consórcio
        'G': 15,  # Status
        'H': 40,  # Fluxo de Pagamento
        'I': 15,  # Vencimento
        'J': 40   # Observações
    }
    
    # Aplicar larguras e alinhamento central para todas as células
    for row in worksheet.iter_rows():
        for cell in row:
            # Aplicar alinhamento central
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
    # Aplicar larguras das colunas
    for coluna, largura in larguras_colunas.items():
        worksheet.column_dimensions[coluna].width = largura
    
    # Salvar as alterações
    workbook.save(caminho_completo)
    print(f"Arquivo salvo em: {caminho_completo}")
    return caminho_completo

def carregar_arquivo_base():
    """Carrega o arquivo base se existir, ou cria um novo se não existir."""
    try:
        # Procura pelo arquivo mais recente na pasta DADOS EXTRAIDOS
        pasta_dados = 'DADOS EXTRAIDOS'
        if not os.path.exists(pasta_dados):
            os.makedirs(pasta_dados)
            
        nome_arquivo = 'extracao_SP_contemplados.xlsx'
        arquivo_base = os.path.join(pasta_dados, nome_arquivo)
        
        if not os.path.exists(arquivo_base):
            # Se não existir arquivo, cria um novo DataFrame vazio
            print("Nenhum arquivo base encontrado. Será criado um novo.")
            df_vazio = pd.DataFrame(columns=['Código', 'Tipo', 'Valor da carta', 'Entrada', 
                                           'Total de Parcelas', 'Consórcio', 'Status', 
                                           'Fluxo de Pagamento', 'Vencimento', 'Observações'])
            return df_vazio
        
        print(f"Carregando arquivo base: {arquivo_base}")
        return pd.read_excel(arquivo_base)
    except Exception as e:
        print(f"Erro ao carregar arquivo base: {str(e)}")
        # Em caso de erro, retorna um DataFrame vazio
        return pd.DataFrame(columns=['Código', 'Tipo', 'Valor da carta', 'Entrada', 
                                   'Total de Parcelas', 'Consórcio', 'Status', 
                                   'Fluxo de Pagamento', 'Vencimento', 'Observações'])

def comparar_e_atualizar_dados(df_novo, df_base):
    """Compara os dados novos com o arquivo base e atualiza apenas o que mudou."""
    # Se o DataFrame base estiver vazio, retorna o novo DataFrame
    if df_base.empty:
        return df_novo
    
    # Cria uma cópia do DataFrame base
    df_atualizado = df_base.copy()
    
    # Remove cotas vendidas do arquivo base
    df_atualizado = df_atualizado[df_atualizado['Status'] != 'Vendida']
    
    # Para cada linha no novo DataFrame
    for _, linha_nova in df_novo.iterrows():
        # Procura a linha correspondente no arquivo base pelo código
        idx = df_atualizado[df_atualizado['Código'] == linha_nova['Código']].index
        
        if len(idx) > 0:
            # Se encontrou, atualiza apenas se houver mudanças
            if df_atualizado.loc[idx[0], 'Status'] != linha_nova['Status'] or \
               df_atualizado.loc[idx[0], 'Valor da carta'] != linha_nova['Valor da carta'] or \
               df_atualizado.loc[idx[0], 'Entrada'] != linha_nova['Entrada'] or \
               df_atualizado.loc[idx[0], 'Total de Parcelas'] != linha_nova['Total de Parcelas'] or \
               df_atualizado.loc[idx[0], 'Consórcio'] != linha_nova['Consórcio'] or \
               df_atualizado.loc[idx[0], 'Fluxo de Pagamento'] != linha_nova['Fluxo de Pagamento']:
                df_atualizado.loc[idx[0]] = linha_nova
        else:
            # Se não encontrou, adiciona a nova linha
            df_atualizado = pd.concat([df_atualizado, pd.DataFrame([linha_nova])], ignore_index=True)
    
    return df_atualizado

def executar_scraping(callback=None):
    """Função principal que executa todo o processo de scraping."""
    def log(message):
        print(message)
        if callback:
            callback(message)
    
    driver = None
    max_tentativas = 3  # Número máximo de tentativas
    
    for tentativa in range(max_tentativas):
        try:
            log(f"\nTentativa {tentativa + 1} de {max_tentativas}")
            log("Iniciando processo de scraping...")
            
            # Carregar arquivo base
            df_base = carregar_arquivo_base()
            
            # URLs específicas
            url_imoveis = "https://www.contempladosp.com.br/cartas-de-credito-contemplada-de-imoveis"
            url_veiculos = "https://www.contempladosp.com.br/cartas-de-credito-contemplada-de-veiculos"
            
            # Configurar e iniciar o driver
            driver = configurar_driver()
            
            # Extrair dados de imóveis
            log("\nAcessando página de imóveis...")
            driver.get(url_imoveis)
            
            # Aguardar carregamento da página
            log("Aguardando carregamento da página de imóveis...")
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.CLASS_NAME, "table"))
            )
            
            # Aguardar um pouco mais para garantir que o JavaScript carregou tudo
            time.sleep(15)
            
            # Obter o HTML da página de imóveis
            log("Extraindo HTML da página de imóveis...")
            html_imoveis = driver.page_source
            soup_imoveis = BeautifulSoup(html_imoveis, 'html.parser')
            dados_imoveis, contador = extrair_dados_tabela(soup_imoveis, 'Imóveis', 1)
            
            # Extrair dados de veículos
            log("\nAcessando página de veículos...")
            driver.get(url_veiculos)
            
            # Aguardar carregamento da página
            log("Aguardando carregamento da página de veículos...")
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.CLASS_NAME, "table"))
            )
            
            # Aguardar um pouco mais para garantir que o JavaScript carregou tudo
            time.sleep(15)
            
            # Obter o HTML da página de veículos
            log("Extraindo HTML da página de veículos...")
            html_veiculos = driver.page_source
            soup_veiculos = BeautifulSoup(html_veiculos, 'html.parser')
            dados_veiculos, _ = extrair_dados_tabela(soup_veiculos, 'Veículos', contador)
            
            if not dados_imoveis and not dados_veiculos:
                log("ERRO: Nenhum dado foi extraído das tabelas!")
                continue
            
            # Criar DataFrame com os novos dados
            df_novo = criar_dataframe(dados_imoveis, dados_veiculos)
            
            # Comparar e atualizar dados
            df_final = comparar_e_atualizar_dados(df_novo, df_base)
            
            # Salvar resultado
            data_atual = datetime.now()
            arquivo_salvo = salvar_excel(df_final, data_atual)
            
            # Gerar estatísticas detalhadas
            log("\nEstatísticas detalhadas:")
            log(f"Total de registros: {len(df_final)}")
            log(f"- Imóveis: {len(dados_imoveis)} registros")
            log(f"- Veículos: {len(dados_veiculos)} registros")
            
            # Estatísticas por status
            status_counts = df_final['Status'].value_counts()
            log("\nStatus dos consórcios:")
            for status, count in status_counts.items():
                log(f"- {status}: {count} registros")
            
            # Estatísticas por consórcio e status
            log("\nStatus por administradora:")
            for consorcio in df_final['Consórcio'].unique():
                df_consorcio = df_final[df_final['Consórcio'] == consorcio]
                status_consorcio = df_consorcio['Status'].value_counts()
                log(f"\n{consorcio}:")
                for status, count in status_consorcio.items():
                    log(f"- {status}: {count} registros")
            
            # Se chegou aqui, deu certo!
            break
            
        except Exception as e:
            log(f"\nERRO durante a execução (tentativa {tentativa + 1}): {str(e)}")
            if tentativa < max_tentativas - 1:
                log("Tentando novamente em 10 segundos...")
                time.sleep(10)
            else:
                log("Número máximo de tentativas atingido.")
                import traceback
                log(traceback.format_exc())
        
        finally:
            if driver:
                log("\nFechando o driver...")
                driver.quit()
    
    if tentativa == max_tentativas - 1:
        log("\nNão foi possível completar o scraping após todas as tentativas.")

def main():
    # Cria pasta para os arquivos se não existir
    pasta_dados = "DADOS EXTRAIDOS"
    if not os.path.exists(pasta_dados):
        os.makedirs(pasta_dados)

    # Nome do arquivo base
    nome_arquivo = 'extracao_SP_contemplados.xlsx'
    arquivo_base = os.path.join(pasta_dados, nome_arquivo)

    print(f"\nTentativa 1 de 3")
    print("Iniciando processo de scraping...")
    
    if os.path.exists(arquivo_base):
        print(f"Carregando arquivo base: {arquivo_base}")
    else:
        print("Nenhum arquivo base encontrado. Será criado um novo.")
    
    print("Configurando o driver do Chrome...")
    driver = configurar_driver()

    try:
        # URLs específicas
        url_imoveis = "https://www.contempladosp.com.br/cartas-de-credito-contemplada-de-imoveis"
        url_veiculos = "https://www.contempladosp.com.br/cartas-de-credito-contemplada-de-veiculos"
        
        # Extrair dados de imóveis
        print("\nAcessando página de imóveis...")
        driver.get(url_imoveis)
        
        # Aguardar carregamento da página
        print("Aguardando carregamento da página de imóveis...")
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.CLASS_NAME, "table"))
        )
        
        # Aguardar um pouco mais para garantir que o JavaScript carregou tudo
        time.sleep(15)
        
        # Obter o HTML da página de imóveis
        print("Extraindo HTML da página de imóveis...")
        html_imoveis = driver.page_source
        soup_imoveis = BeautifulSoup(html_imoveis, 'html.parser')
        dados_imoveis, contador = extrair_dados_tabela(soup_imoveis, 'Imóveis', 1)
        
        # Extrair dados de veículos
        print("\nAcessando página de veículos...")
        driver.get(url_veiculos)
        
        # Aguardar carregamento da página
        print("Aguardando carregamento da página de veículos...")
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.CLASS_NAME, "table"))
        )
        
        # Aguardar um pouco mais para garantir que o JavaScript carregou tudo
        time.sleep(15)
        
        # Obter o HTML da página de veículos
        print("Extraindo HTML da página de veículos...")
        html_veiculos = driver.page_source
        soup_veiculos = BeautifulSoup(html_veiculos, 'html.parser')
        dados_veiculos, _ = extrair_dados_tabela(soup_veiculos, 'Veículos', contador)
        
        if not dados_imoveis and not dados_veiculos:
            print("ERRO: Nenhum dado foi extraído das tabelas!")
            return
        
        # Criar DataFrame com os novos dados
        df_novo = criar_dataframe(dados_imoveis, dados_veiculos)
        
        # Carregar dados existentes
        if os.path.exists(arquivo_base):
            df_existente = pd.read_excel(arquivo_base)
            print(f"Arquivo base encontrado com {len(df_existente)} registros")
        else:
            df_existente = pd.DataFrame()
            print("Criando novo arquivo base")
        
        # Comparar e atualizar dados
        df_atualizado = comparar_e_atualizar_dados(df_novo, df_existente)
        
        # Salvar arquivo usando a função salvar_excel
        data_atual = datetime.now()
        arquivo_salvo = salvar_excel(df_atualizado, data_atual)
        print(f"Arquivo atualizado em: {arquivo_salvo}")
        
        # Gerar estatísticas detalhadas
        print("\nEstatísticas detalhadas:")
        print(f"Total de registros: {len(df_atualizado)}")
        print(f"- Imóveis: {len(dados_imoveis)} registros")
        print(f"- Veículos: {len(dados_veiculos)} registros")
        
        # Estatísticas por status
        status_counts = df_atualizado['Status'].value_counts()
        print("\nStatus dos consórcios:")
        for status, count in status_counts.items():
            print(f"- {status}: {count} registros")
        
        # Estatísticas por consórcio e status
        print("\nStatus por administradora:")
        for consorcio in df_atualizado['Consórcio'].unique():
            df_consorcio = df_atualizado[df_atualizado['Consórcio'] == consorcio]
            status_consorcio = df_consorcio['Status'].value_counts()
            print(f"\n{consorcio}:")
            for status, count in status_consorcio.items():
                print(f"- {status}: {count} registros")
        
    except Exception as e:
        print(f"Erro durante a execução: {str(e)}")
        import traceback
        print(traceback.format_exc())
    
    finally:
        if driver:
            print("\nFechando o driver...")
            driver.quit()

if __name__ == "__main__":
    main() 